"""
MSA publication CIN report.
Columns: MSA, County, Host, [10 CIN categories]
Only includes sources assigned to one of the 8 defined MSAs.
Output: ciep_march_msa_publication_cin.csv
"""

import csv
import os
import re

from openpyxl import load_workbook

PUB_CSV = os.path.join(os.path.dirname(__file__), "ciep_march_publication_report.csv")
OUTPUT  = os.path.join(os.path.dirname(__file__), "ciep_march_msa_publication_cin.csv")
WORKBOOK = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Documents/Mizzou/CIEP/"
    "MO Report/ciep_march_base_metrics.xlsx"
)

CIN_COLS = [
    "Civic Life", "Civic information", "Economic Development", "Education",
    "Emergencies and Public Safety", "Environment and Planning", "Health",
    "Political life", "Sports", "Transportation Systems",
]

NON_MSA = "Non-MSA / Unknown"


def normalize_host(host: str) -> str:
    host = (host or "").strip().lower()
    host = re.sub(r"^www\.", "", host)
    return host.split("/")[0]


def normalize_county(raw: str) -> str:
    raw = (raw or "").strip().lower()
    raw = raw.replace(".", "")
    raw = re.sub(r"\bcounty\b", "", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    aliases = {
        "cape giradeau": "cape girardeau",
        "st louis city": "st louis",
    }
    raw = aliases.get(raw, raw)
    return raw


# Load publication profile rows from workbook and derive MSA from publication_county.
county_to_msa = {}
for msa, counties in {
    "St. Louis, MO-IL": [
        "St. Louis", "St. Charles", "Jefferson", "Franklin", "Lincoln", "Warren",
    ],
    "Kansas City, MO-KS": [
        "Jackson", "Clay", "Platte", "Cass", "Ray", "Lafayette", "Caldwell", "Wyandotte",
    ],
    "Springfield": ["Greene", "Christian", "Webster", "Dallas"],
    "Columbia": ["Boone"],
    "Joplin": ["Jasper", "Newton"],
    "Jefferson City": ["Cole", "Callaway", "Moniteau", "Osage"],
    "St. Joseph, MO-KS": ["Buchanan", "DeKalb", "Andrew"],
    "Cape Girardeau, MO-IL": ["Cape Girardeau", "Bollinger"],
}.items():
    for county in counties:
        key = normalize_county(county)
        county_to_msa[key] = msa
        county_to_msa[key.replace(" county", "").strip()] = msa

wb = load_workbook(WORKBOOK, data_only=True)
ws = wb["Publication Profile"]

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
host_col = headers.index("host") + 1
county_col = headers.index("publication_county") + 1 if "publication_county" in headers else None
if county_col is None:
    raise RuntimeError("publication_county column not found in Publication Profile tab")

msa_map = {}
for r in range(2, ws.max_row + 1):
    host = normalize_host(str(ws.cell(r, host_col).value or ""))
    county_raw = str(ws.cell(r, county_col).value or "").strip()
    if not host or not county_raw:
        continue

    msa = None
    for part in re.split(r"\s+and\s+|/|,|&", county_raw, flags=re.IGNORECASE):
        norm = normalize_county(part)
        if norm in county_to_msa:
            msa = county_to_msa[norm]
            break
    if not msa:
        msa = NON_MSA

    if msa != NON_MSA:
        msa_map[host] = {"msa": msa, "county": county_raw}

# Load publication CIN counts: host → {col: count}
pub_cin = {}
with open(PUB_CSV, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        pub_cin[normalize_host(row["publication"])] = row

# Build output rows (MSA sources only)
rows = []
for host, info in msa_map.items():
    cin = pub_cin.get(host, {})
    row = {
        "MSA":            info["msa"],
        "County":         info["county"],
        "Host":           host,
        "total_articles": cin.get("total_articles", 0),
    }
    for col in CIN_COLS:
        row[col] = cin.get(col, 0)
    rows.append(row)

# Sort by MSA order, then county, then host
MSA_ORDER = [
    "St. Louis, MO-IL", "Kansas City, MO-KS", "Springfield",
    "Columbia", "Joplin", "Jefferson City",
    "St. Joseph, MO-KS", "Cape Girardeau, MO-IL",
]
rows.sort(key=lambda r: (MSA_ORDER.index(r["MSA"]) if r["MSA"] in MSA_ORDER else 99,
                          r["County"], r["Host"]))

fieldnames = ["MSA", "County", "Host", "total_articles"] + CIN_COLS
with open(OUTPUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

print(f"Saved: {OUTPUT}")
print(f"Rows: {len(rows)}")
