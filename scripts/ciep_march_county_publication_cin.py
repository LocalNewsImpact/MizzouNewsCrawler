"""County publication CIN report.

Aggregates the March publication-level CIN report by county using the
`publication_county` column from the Publication Profile workbook tab.

Output: ciep_march_county_publication_cin.csv
"""

import csv
import os
import re
from collections import defaultdict

from openpyxl import load_workbook

PUB_CSV = os.path.join(os.path.dirname(__file__), "ciep_march_publication_report.csv")
OUTPUT = os.path.join(os.path.dirname(__file__), "ciep_march_county_publication_cin.csv")
WORKBOOK = os.path.expanduser(
    "~/Library/Mobile Documents/com~apple~CloudDocs/Documents/Mizzou/CIEP/"
    "MO Report/ciep_march_base_metrics.xlsx"
)

CIN_COLS = [
    "Civic Life", "Civic information", "Economic Development", "Education",
    "Emergencies and Public Safety", "Environment and Planning", "Health",
    "Political life", "Sports", "Transportation Systems",
]


def normalize_host(host: str) -> str:
    host = (host or "").strip().lower()
    host = re.sub(r"^www\.", "", host)
    return host.split("/")[0]


def normalize_county(raw: str) -> str:
    raw = (raw or "").strip().lower()
    raw = raw.replace(".", "")
    raw = re.sub(r"\bcounty\b", "", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    aliases = {
        "cape giradeau": "cape girardeau",
        "st louis city": "st louis",
    }
    return aliases.get(raw, raw)


def split_counties(raw: str) -> list[str]:
    raw = (raw or "").strip()
    if not raw:
        return []

    counties = []
    for part in re.split(r"\s+and\s+|/|,|&", raw, flags=re.IGNORECASE):
        county = normalize_county(part)
        if county and county not in counties:
            counties.append(county)
    return counties


wb = load_workbook(WORKBOOK, data_only=True)
ws = wb["Publication Profile"]

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
host_col = headers.index("host") + 1
county_col = headers.index("publication_county") + 1 if "publication_county" in headers else None
if county_col is None:
    raise RuntimeError("publication_county column not found in Publication Profile tab")

host_to_counties = {}
for r in range(2, ws.max_row + 1):
    host = normalize_host(str(ws.cell(r, host_col).value or ""))
    county_raw = str(ws.cell(r, county_col).value or "").strip()
    if not host or not county_raw:
        continue
    counties = split_counties(county_raw)
    if counties:
        host_to_counties[host] = counties


county_hosts = defaultdict(set)
county_totals = defaultdict(lambda: {"total_articles": 0, **dict.fromkeys(CIN_COLS, 0)})

with open(PUB_CSV, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        host = normalize_host(row["publication"])
        counties = host_to_counties.get(host, [])
        if not counties:
            continue

        for county in counties:
            county_hosts[county].add(host)
            try:
                county_totals[county]["total_articles"] += int(float(row.get("total_articles", 0) or 0))
            except ValueError:
                pass

            for col in CIN_COLS:
                try:
                    county_totals[county][col] += int(float(row.get(col, 0) or 0))
                except ValueError:
                    pass


rows = []
for county, totals in county_totals.items():
    row = {
        "County": county,
        "publications": len(county_hosts[county]),
        "total_articles": totals["total_articles"],
    }
    for col in CIN_COLS:
        row[col] = totals[col]
    rows.append(row)

rows.sort(key=lambda r: (-r["total_articles"], r["County"]))

fieldnames = ["County", "publications", "total_articles"] + CIN_COLS
with open(OUTPUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

print(f"Saved: {OUTPUT}")
print(f"Rows: {len(rows)}")
